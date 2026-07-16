"""
Builds the BDR-facing Excel workbook.

Sheet 1 - Enriched Data: every original scrape field + all enrichment
    columns + Tier classification, for every row (qualified AND
    disqualified - nothing is deleted).
Sheet 2 - Ranked Lead List: sorted Tier 1 -> Tier 2 -> Tier 3 ->
    Disqualified, with an auto-filter enabled so disqualified rows
    stay visible and filterable rather than hidden.
Sheet 3 - Cluster Analysis: one row per cluster with its size, Tier
    1/2/3 counts, and which named shops it contains.
Real Shop Demo - the separate, real-shop enrichment worked example
    (see enrichment.py), clearly labelled as a worked example and never
    blended into the dummy-scrape sheets above.

Scale: a fully-styled Excel sheet with tens of thousands of rows is
impractical to build and to browse. If `full_df` is larger than
`max_styled_rows`, the complete result is written to CSV alongside the
workbook and the styled Sheet 1 / Sheet 2 are capped to a bounded top
slice (Tier 1 + Tier 2 rows first, then Tier 3, up to the cap) - see
the README for the exact threshold and rationale.
"""
import os

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

HEADER_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
SECTION_FONT = Font(bold=True, size=12)
WRAP_ALIGNMENT = Alignment(wrap_text=True, vertical="top")
TOP_ALIGNMENT = Alignment(vertical="top")
POSSIBLE_DUPLICATE_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

# Columns whose content is long free text and should wrap rather than
# be squeezed onto one line.
WRAP_COLUMNS = {
    "qualification_reason", "tier_reasoning", "top_review",
    "research_notes", "price_band", "notes", "SUGGESTED_MESSAGE",
    "MESSAGE_LOGIC", "_ELIGIBILITY_REASON", "priority_reasoning",
}

DEFAULT_MAX_STYLED_ROWS = 2000


def _write_sheet(wb: Workbook, sheet_name: str, df: pd.DataFrame, freeze_first_col: bool = False,
                  highlight_col: str = None, highlight_contains: str = None, note: str = None):
    """highlight_col/highlight_contains: if given, any row whose
    highlight_col value contains highlight_contains gets a fill colour
    (see POSSIBLE_DUPLICATE_FILL) - used for the Rule 2 "flagged, not
    merged" duplicate rows so they're visually distinguishable, not just
    text-flagged.

    note: an optional italic caption row written above the header (e.g.
    a colour-legend explanation) - freeze panes and auto-filter are
    computed to account for the extra row rather than bolted on
    afterwards, which would leave them pointing at the wrong rows."""
    ws: Worksheet = wb.create_sheet(sheet_name)
    header_row = 1
    if note:
        ws.append([note])
        ws["A1"].font = Font(italic=True, size=9)
        header_row = 2

    if df.empty:
        ws.append(["(no rows)"])
        return ws

    columns = list(df.columns)
    ws.append(columns)
    for cell in ws[header_row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = TOP_ALIGNMENT

    highlight_idx = columns.index(highlight_col) + 1 if highlight_col in columns else None
    for row in df.itertuples(index=False):
        ws.append(list(row))
        if highlight_idx is not None:
            value = str(row[highlight_idx - 1])
            if highlight_contains in value:
                for cell in ws[ws.max_row]:
                    cell.fill = POSSIBLE_DUPLICATE_FILL

    for col_idx, col_name in enumerate(columns, start=1):
        letter = get_column_letter(col_idx)
        wrap = col_name in WRAP_COLUMNS
        width = 40 if wrap else min(max(len(str(col_name)) + 2, 12), 30)
        ws.column_dimensions[letter].width = width
        data_rows = ws[letter][header_row:]
        for cell in data_rows:
            cell.alignment = WRAP_ALIGNMENT if wrap else TOP_ALIGNMENT

    freeze_col = "B" if freeze_first_col else "A"
    ws.freeze_panes = f"{freeze_col}{header_row + 1}"
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(columns))}{ws.max_row}"
    return ws


def _cluster_analysis(full_df: pd.DataFrame) -> pd.DataFrame:
    clustered = full_df[full_df["cluster_id"].notna()].copy() if "cluster_id" in full_df.columns else pd.DataFrame()
    if clustered.empty:
        return pd.DataFrame(columns=[
            "cluster_id", "cluster_size", "tier_1_count", "tier_2_count",
            "tier_3_count", "shops_in_cluster",
        ])

    rows = []
    for cluster_id, group in clustered.groupby("cluster_id"):
        tier_counts = group["tier"].value_counts()
        rows.append({
            "cluster_id": cluster_id,
            "cluster_size": len(group),
            "tier_1_count": int(tier_counts.get("Tier 1", 0)),
            "tier_2_count": int(tier_counts.get("Tier 2", 0)),
            "tier_3_count": int(tier_counts.get("Tier 3", 0)),
            "shops_in_cluster": "; ".join(group["place_name"].astype(str)),
        })
    out = pd.DataFrame(rows).sort_values("cluster_size", ascending=False).reset_index(drop=True)
    return out


def _ranked(full_df: pd.DataFrame) -> pd.DataFrame:
    tier_order = {"Tier 1": 0, "Tier 2": 1, "Tier 3": 2, "Disqualified": 3}
    df = full_df.copy()
    df["_sort"] = df["tier"].map(tier_order).fillna(3)
    return df.sort_values("_sort").drop(columns="_sort").reset_index(drop=True)


def _cap_for_styling(df: pd.DataFrame, max_rows: int) -> pd.DataFrame:
    if len(df) <= max_rows:
        return df
    ranked = _ranked(df)
    return ranked.head(max_rows)


def build_workbook(full_df: pd.DataFrame, output_path: str,
                    real_shop_df: pd.DataFrame = None,
                    max_styled_rows: int = DEFAULT_MAX_STYLED_ROWS) -> dict:
    """full_df must already carry qualifies/qualification_reason, tier/
    tier_reasoning, and (for qualified rows) cluster_id/cluster_size -
    i.e. it's the merged qualified+disqualified, tiered, clustered
    result the CLI assembles before calling this.

    real_shop_df, if given, must already be qualified+tiered the same
    way (see enrichment.py + cli.py) and is written to its own sheet,
    clearly separate from the dummy-scrape sheets.

    Returns a dict describing what was written (paths + row counts) so
    the CLI can report it to the user."""
    result = {"workbook": output_path}

    oversized = len(full_df) > max_styled_rows
    if oversized:
        csv_path = os.path.splitext(output_path)[0] + "_full.csv"
        full_df.to_csv(csv_path, index=False)
        result["full_csv"] = csv_path
        result["full_csv_rows"] = len(full_df)
        styled_df = _cap_for_styling(full_df, max_styled_rows)
        result["excel_capped"] = True
        result["excel_rows"] = len(styled_df)
    else:
        styled_df = full_df
        result["excel_capped"] = False
        result["excel_rows"] = len(styled_df)

    wb = Workbook()
    wb.remove(wb.active)

    _write_sheet(wb, "Enriched Data", styled_df, freeze_first_col=True)
    _write_sheet(wb, "Ranked Lead List", _ranked(styled_df), freeze_first_col=True)
    _write_sheet(wb, "Cluster Analysis", _cluster_analysis(styled_df))

    if real_shop_df is not None and not real_shop_df.empty:
        _write_sheet(wb, "Real Shop Demo (Manchester)", _ranked(real_shop_df), freeze_first_col=True)
        _write_sheet(wb, "Real Shop Demo - Clusters", _cluster_analysis(real_shop_df))
        result["real_shop_rows"] = len(real_shop_df)

    os.makedirs(os.path.dirname(os.path.abspath(output_path)) or ".", exist_ok=True)
    wb.save(output_path)
    return result


# =============================================================================
# Part 2 - CRM cleaning + outreach workbook
# =============================================================================
def _write_log_sheet(wb: Workbook, log: dict, qa: list, flagged_pairs: list):
    """Cleaning Log sheet: counts, QA results, and the documented
    assumptions the brief asks to log rather than silently apply."""
    ws: Worksheet = wb.create_sheet("Cleaning Log")
    ws.column_dimensions["A"].width = 46
    ws.column_dimensions["B"].width = 70

    def section(title):
        ws.append([title])
        ws[f"A{ws.max_row}"].font = SECTION_FONT
        ws.append([])

    def kv(key, value):
        ws.append([key, value])
        ws[f"B{ws.max_row}"].alignment = WRAP_ALIGNMENT

    section("Row counts")
    kv("Starting rows", log.get("starting_rows"))
    kv("Removed as 'No Fit' (Rule 1)", log.get("no_fit_removed"))
    kv("Duplicate rows consolidated (Rule 2)", log.get("duplicate_rows_consolidated"))
    kv("Unique business groups after dedup", log.get("unique_business_groups"))
    kv("Disqualified by category/charity-name qualification (Rule 9.5)", log.get("disqualified_by_rule_9_5"))
    kv("Qualified leads (final Cleaned Dataset row count)", log.get("qualified_leads"))
    ws.append([])

    section("Rule 5 - Non-UK handling")
    if "non_uk_flagged_not_removed" in log:
        kv("Non-UK leads FLAGGED (not removed - see README deviation note)", log["non_uk_flagged_not_removed"])
    else:
        kv("Non-UK leads REMOVED (--remove-non-uk passed, matching the original brief literally)", log.get("non_uk_removed"))
    kv("Rows with no country data at all (flagged 'Unverified Location', never assumed UK)", log.get("unverified_location_count"))
    ws.append([])

    section("Rule 2 - Duplicate candidates flagged, not merged")
    kv("Pairs flagged for manual review (city/country conflict guard fired)", log.get("possible_duplicate_pairs_flagged"))
    for a, b, key, city_a, city_b, country_a, country_b in flagged_pairs:
        kv(f"  {a} vs {b} (matched on {key})", f"{city_a}, {country_a}  vs  {city_b}, {country_b}")
    ws.append([])

    section("Rule 7 - Contact information completion")
    kv("Fields found via genuine research", log.get("contact_fields_found_via_research"))
    kv("Fields defaulted to 'unknown' (business names in this dataset are synthetic/placeholder - "
       "see README)", log.get("contact_fields_defaulted_to_unknown"))
    ws.append([])

    section("Rule 13 - QA checks")
    kv("Checks passed", f"{log.get('qa_checks_passed')} / {log.get('qa_checks_total')}")
    for name, passed, detail in qa:
        kv(f"  [{'PASS' if passed else 'FAIL'}] {name}", detail)
    ws.append([])

    section("Documented assumptions and resolutions")
    kv("Rule 11 contactability_score contradiction",
       "The brief's own table has '2 = Missing three' and '1 = Only one contact method' both describing "
       "exactly 1-of-4 contact methods present, with different scores. Resolved by trusting the more "
       "structured 'missing N' framing (a clean monotonic scale: 0 methods=0, 1=2, 2=3, 3=4, 4=5), since "
       "5 of 6 rows in the brief's table agree on that reading.")
    kv("Rule 6 bare-date year inference",
       "Dates with no year (e.g. '20 Jun') are assumed to fall near the cleaning run's anchor date, rolled "
       "back a year if that would place them implausibly in the future - a genuine assumption, not a "
       "certainty, applied via parse_date_flex().")
    kv("Rule 6 extension: last_purchase_date 'never'",
       "The brief's Special Rule only names last_contact_date for the 'never' fallback; extended to "
       "last_purchase_date too, since most leads simply haven't purchased yet, which 'never' states "
       "accurately - the generic 'unknown' fallback would wrongly imply doubt.")
    kv("Rule 5 flag-vs-remove choice",
       "The original brief says REMOVE non-UK leads; flag_non_uk() defaults to remove=False (flag only) "
       "instead, per real client preference - remove=True is a one-line change away, not a rebuild.")
    kv("Date parser bug fix (found running this pipeline)",
       "parse_date_flex()'s dayfirst=True was misreading already-year-first dates like '2026/04/10' as "
       "4 October instead of 10 April, producing last-contact dates in the future. Fixed to only apply "
       "dayfirst semantics when the string does not already start with a 4-digit year; does not change "
       "the pinned qualified/duplicate-group counts (date parsing doesn't affect qualification or dedup "
       "group membership, only the merged date value and tie-breaking).")


def _write_summary_sheet(wb: Workbook, qualified_df: pd.DataFrame, city_spend_tier_breakdown: pd.DataFrame = None):
    """Channel & CRM Summary: several small tables in one sheet, since
    none of them individually need their own tab."""
    ws: Worksheet = wb.create_sheet("Channel & CRM Summary")
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 16
    for letter in ("C", "D", "E", "F"):
        ws.column_dimensions[letter].width = 12

    def table(title, series_or_df):
        ws.append([title])
        ws[f"A{ws.max_row}"].font = SECTION_FONT
        if isinstance(series_or_df, pd.Series):
            for idx, val in series_or_df.items():
                ws.append([str(idx), val])
        elif isinstance(series_or_df, pd.DataFrame):
            ws.append(list(series_or_df.columns))
            for cell in ws[ws.max_row]:
                cell.font = Font(bold=True)
            for row in series_or_df.itertuples(index=False):
                ws.append(list(row))
        ws.append([])

    table("Store type (Physical Store vs Online Retailer)", qualified_df["store_type"].value_counts())
    table("Leads by city", qualified_df["city"].value_counts().head(20))
    table("Leads by pipeline_status", qualified_df["pipeline_status"].value_counts())
    table("Leads by stage_clean", qualified_df["stage_clean"].value_counts())

    active_customer = qualified_df["pipeline_status"].eq("Customer").sum()
    churned = qualified_df["pipeline_status"].eq("Churned").sum()
    ws.append(["Active vs Churned customers"])
    ws[f"A{ws.max_row}"].font = SECTION_FONT
    ws.append(["Active Customer (pipeline_status=Customer)", int(active_customer)])
    ws.append(["Churned (pipeline_status=Churned)", int(churned)])
    ws.append([])

    table("Contactability score distribution", qualified_df["contactability_score"].value_counts().sort_index())
    table("Data quality flag counts", qualified_df["data_quality_flag"].value_counts())

    if "spend_tier" in qualified_df.columns:
        tier_order = ["Tier 1", "Tier 2", "Tier 3", "Unknown"]
        counts = qualified_df["spend_tier"].value_counts().reindex(tier_order).dropna()
        table("Leads by spend_tier (Tier 1: GBP5k+/mo, Tier 2: GBP2k-4,999/mo, Tier 3: under GBP2k/mo)", counts)

    if city_spend_tier_breakdown is not None and not city_spend_tier_breakdown.empty:
        table("Lead count by city and spend_tier (sorted by total leads, descending)", city_spend_tier_breakdown)


_COLOUR_LEGEND_NOTE = (
    "Row shading: yellow = Possible Duplicate - Unverified (Rule 2 flagged this row as a candidate "
    "match with a city/country conflict - kept separate, not merged; verify by hand before treating "
    "as the same business)."
)


_CITY_PRIORITY_NOTE = (
    "Ranked by priority_score (density 25% + spend potential 35% + pipeline warmth 25% + win rate 15%, "
    "each normalised 0-1 across cities). priority_reasoning spells out the real numbers behind every "
    "score - never trust the score alone without reading it. Physical Store leads with a known city only; "
    "online resellers are reached by DM, not a city visit, and leads with no city can't inform this."
)


def build_crm_workbook(qualified_df: pd.DataFrame, disqualified_df: pd.DataFrame,
                        log: dict, qa: list, flagged_pairs: list, output_path: str,
                        city_spend_tier_breakdown: pd.DataFrame = None,
                        city_prioritisation: pd.DataFrame = None) -> dict:
    """Builds the Part 2 workbook: Cleaned Dataset, Disqualified Leads,
    Cleaning Log, Channel & CRM Summary, City Prioritisation.
    qualified_df should already carry the Part B outreach columns
    (OUTREACH_TYPE, SUGGESTED_MESSAGE, MESSAGE_LOGIC,
    PERSONALISATION_ANGLE, RECOMMENDED_FOLLOW_UP_DATE) if
    generate-outreach has been run; if not, the Cleaned Dataset sheet
    simply omits them. city_spend_tier_breakdown (see
    crm_cleaning.city_spend_tier_breakdown()), if given, is rendered as
    a table on the Channel & CRM Summary sheet. city_prioritisation
    (see city_priority.city_prioritisation()), if given, gets its own
    sheet - it's the answer to "which city should we go after next".
    """
    wb = Workbook()
    wb.remove(wb.active)

    _write_sheet(wb, "Cleaned Dataset", qualified_df, freeze_first_col=True,
                 highlight_col="data_quality_flag", highlight_contains="Possible Duplicate - Unverified",
                 note=_COLOUR_LEGEND_NOTE)
    _write_sheet(wb, "Disqualified Leads", disqualified_df, freeze_first_col=True)
    _write_log_sheet(wb, log, qa, flagged_pairs)
    _write_summary_sheet(wb, qualified_df, city_spend_tier_breakdown=city_spend_tier_breakdown)

    if city_prioritisation is not None and not city_prioritisation.empty:
        _write_sheet(wb, "City Prioritisation", city_prioritisation, freeze_first_col=True,
                     note=_CITY_PRIORITY_NOTE)

    os.makedirs(os.path.dirname(os.path.abspath(output_path)) or ".", exist_ok=True)
    wb.save(output_path)
    return {"workbook": output_path, "qualified_rows": len(qualified_df), "disqualified_rows": len(disqualified_df)}
